#!/usr/bin/env python
# coding: utf-8


# Importing required modules from openpyxl package and also importing styles required for formatting the excel cells
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Fill,Font,Color,colors


#Loading the excel workbooks
wb_final= Workbook()
wb_old=load_workbook('Excel_old.xlsx')
wb_new=load_workbook('Excel_new.xlsx')


#reading the sheets from the workbook
sh1= wb_old.active
sh2= wb_new.active
sh3= wb_final.active

#adding the header files to the final worksheet
for i in range(1,sh2.max_column+1):
    sh3.cell(1,i).value=sh1.cell(1,i).value
wb_final.save('Excel_final.xlsx')

#Calculating the difference and add it to the new workbook and then highlight the changes by colouring it
for i in range(2,sh2.max_row+1):
    for j in range(1, sh2.max_column+1):
        if type(sh1.cell(row=i,column=j).value)==str and type(sh2.cell(row=i,column=j).value)==str :
            sh3.cell(row=i, column=j).value = sh2.cell(row=i,column=j).value
        else:
            sh3.cell(row=i,column=j).value = sh1.cell(row=i,column=j).value-sh2.cell(row=i,column=j).value
            if sh3.cell(row=i,column=j).value!=0:
                    sh3.cell(row=i, column=j).font = Font(color=colors.RED)
wb_final.save('Excel_final.xlsx')
